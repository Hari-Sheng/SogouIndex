# encoding=gbk
import urllib2, re, string, time, datetime, os, sys, openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook



reload(sys)
sys.setdefaultencoding('gbk')
CurrentDir = os.getcwd()

def search(name, key, dir):
    CurrentDir = dir
    keyword = key

    url = "http://index.sogou.com/sidx?type=0&query=" + keyword.decode("gbk").encode("gb2312") +"&newstype=-1"
    req = urllib2.urlopen(url)
    # url2 = req.geturl()
    # req = urllib2.Request(url2)
    # doc = urllib2.urlopen(req)

    doc = req.read()
    # print doc

    PeriodMatch = re.search(ur'"\d{4}\W\d{1,2}\W\d{1,2}\|\d{4}\W\d{1,2}\W\d{1,2}"',doc)
    UserIndexMatch = re.search(ur'"userIndexes":"((\d)+,)+(\d)+',doc)
    MediaIndexMatch = re.search(ur"tmp.mediaIndexes =  '((\d)+,)+(\d)+'",doc)
    # print PeriodMatch.group()
    # print UserIndexMatch.group()
    # print UserIndexMatch.group()[15:]
    # print MediaIndexMatch.group()
    # print MediaIndexMatch.group()[21:-1]
    if UserIndexMatch == None or MediaIndexMatch == None:
        print ("|-------------------------------------------------------------------------")
        print ("|Keyword: " + keyword)
        print ("|Sorry, but this keyword hasn't been included by SogouIndex.")

    else:
        UserIndex = UserIndexMatch.group()[15:].split(",")
        MediaIndex = MediaIndexMatch.group()[21:-1].split(",")
        # print UserIndex
        # print MediaIndex
        # print len(UserIndex)
        # print len(MediaIndex)

        print ("|Data received.")


        # print PeriodMatch.group()[1:11]
        Start = time.mktime(time.strptime(PeriodMatch.group()[1:11],'%Y-%m-%d'))
        # print Start
        # print PeriodMatch.group()[12:-1]
        End = time.mktime(time.strptime(PeriodMatch.group()[12:-1],'%Y-%m-%d'))
        # print End
        DayLength = (End-Start)/3600/24 + 1

        if len(UserIndex)==DayLength and len(MediaIndex)==DayLength:
            print ("|Check: The data length is correct.")
        else:
            print ("|Check: Error in data length.")

        Datetime = datetime.datetime.strptime(PeriodMatch.group()[1:11],'%Y-%m-%d')
        # print StartDatetime
        RecordDate = []
        for i in range(0, int(DayLength)):
            RecordDate.append(Datetime.strftime('%Y-%m-%d'))
            Datetime = Datetime + datetime.timedelta(days=1)

        k = 0
        # print RecordDate
        if os.path.isfile(CurrentDir+'\\'+name+'.xlsx') and (keyword!=name):
            wb = load_workbook(filename = name +".xlsx")
            ws = wb.get_sheet_by_name("SogouIndex")

            k = ws.get_highest_column()
            # print(CurrentDir+'\\'+name+'.xlsx')
            # wb = load_workbook
            # print WorkbookTemp.get_sheet(0).cell(0,0).value
            # Workbook = copy(WorkbookTemp)
            # Worksheet = Workbook.get_sheet(0)
            #print Worksheet.cell(0,0),value
            ws.cell(row=1,column=1+k).value=keyword.decode('gbk')
            ws.cell(row=1,column=2+k).value=keyword.decode('gbk')
            ws.cell(row=2,column=1+k).value='UserIndex'
            ws.cell(row=2,column=2+k).value='MediaIndex'

            for i in range(0+k,2+k):
                for j in range(2,int(DayLength)+2):
                    ws.cell(row=j+1,column=1+k).value=int(UserIndex[j-2])
                    ws.cell(row=j+1,column=2+k).value=int(MediaIndex[j-2])
        else:
            wb = Workbook( )
            ws = wb.get_sheet_by_name("Sheet")
            ws.title = 'SogouIndex'
            k=0
            ws.cell(row=1,column=2+k).value=keyword.decode('gbk')
            ws.cell(row=1,column=3+k).value=keyword.decode('gbk')
            ws.cell(row=2,column=1+k).value='Date'
            ws.cell(row=2,column=2+k).value='UserIndex'
            ws.cell(row=2,column=3+k).value='MediaIndex'

            for i in range(0+k,2+k):
                for j in range(2,int(DayLength)+2):
                    ws.cell(row=j+1,column=1+k).value='%s'%RecordDate[j-2]
                    ws.cell(row=j+1,column=2+k).value=int(UserIndex[j-2])
                    ws.cell(row=j+1,column=3+k).value=int(MediaIndex[j-2])


        wb.save(CurrentDir+'\\'+name+'.xlsx')

        print ("|Data saved")
        print ("|-------------------------------------------------------------------------")
        print ("|Keyword: " + keyword)
        if (keyword!=name):
            print ("|Saving Structure: multi-keywords")
        else:
            print ("|Saving Structure: single-keywords")
        print ("|Time span: "+ str(int(DayLength)) +' days')
        print ("|Excel File saved at "+CurrentDir+'\\'+name+'.xlsx')

if __name__=="__main__":
    isLocal= " "
    while not (isLocal=="Y" or isLocal=="N"):
        print isLocal
        isLocal = raw_input("Read from local files?(Y/N): ")

    if isLocal=="Y":
        print ("=========================================================================")
        KeywordList = raw_input("Input the name of file: ")
        while not os.path.isfile(KeywordList+".txt"):
            KeywordList = raw_input("File ["+KeywordList+".txt] does not exist, please input the name of file again:")
        if os.path.isfile(KeywordList+'.xlsx'):
            os.remove(KeywordList+'.xlsx')
        KeywordText = open(KeywordList+".txt").read()
        keywords = re.split('[\s]*[,]*[\s]*[\r]*[\n]*[\t]*',KeywordText)

        isSingleFile = raw_input("Save in a single file?(Y/N): ")
        if isSingleFile =="Y":
        #print re.split('[,]*[\s]*[\r]*[\n]*[\t]*',KeywordText)
        #KeywordText = KeywordText.replace('\r',' ').replace('\n', ' ').replace('\t', ' ').strip()
        #print (KeywordText)

            # print keywords
            for element in keywords:
                search(KeywordList, element, os.getcwd())
        else:
            for element in keywords:
                search(element, element, os.getcwd())

            keyword = raw_input("Press Enter to exit...")
    else:
        while 1:
            print ("=========================================================================")
            keyword = raw_input("Input the keyword (Press Enter to exit): ")
            if len(keyword) == 0:
                break

            # try to merge multi-variable into one excel file
            # elif keyword.find(",")!=-1:
            #     keywords = keyword.split(",")
            #    # os.remove(keywords[0]+'.xls')
            #     for element in keywords:
            #         search(keywords[0], element,os.getcwd())
            # elif keyword.find(" ")!=-1:
            #     keywords = keyword.split(" ")
            #     print keywords
            #     for element in keywords:
            #         search(keywords[0], element, os.getcwd())
            else:
                search(keyword, keyword,os.getcwd())